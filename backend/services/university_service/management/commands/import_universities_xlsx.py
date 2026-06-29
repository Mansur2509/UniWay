"""Import the university XLSX dataset into the catalog.

Usage:
    python manage.py import_universities_xlsx "backend/data/universities/Universities Data.xlsx"

The command is idempotent: universities are matched by a slug derived from the
name (with any trailing "(...)" stripped), so repeated runs upsert instead of
duplicating, and rows that overlap the curated seed catalog enrich the existing
record. See services/university_service/xlsx_import.py for the parsing rules and
docs/DATA_SOURCES.md for the data-quality policy.
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from services.university_service.xlsx_import import (
    EXPECTED_HEADERS,
    VERIFICATION_CHOICES,
    import_rows,
)

SHEET_NAME = "Database"


class Command(BaseCommand):
    help = "Import universities from the dataset XLSX into the catalog (idempotent upsert)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx workbook")
        parser.add_argument(
            "--sheet", default=SHEET_NAME, help=f"Worksheet name (default: {SHEET_NAME})"
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report without writing to the database.",
        )
        parser.add_argument(
            "--replace-existing",
            action="store_true",
            help="Overwrite existing values instead of only filling blanks.",
        )
        parser.add_argument(
            "--include-questionable-stats",
            action="store_true",
            help="Store placeholder-looking SAT values (marked estimated) instead of dropping them.",
        )
        parser.add_argument(
            "--source-label",
            default="Universities Data XLSX",
            help="Label used for created data-source rows.",
        )
        parser.add_argument(
            "--default-verification",
            default="partial",
            choices=sorted(VERIFICATION_CHOICES),
            help="Verification status for sourced fields (default: partial).",
        )
        parser.add_argument(
            "--report",
            default="",
            help="Write the JSON import report to this path (default: alongside the workbook).",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guard
            raise CommandError(
                "openpyxl is required. Install it with: pip install openpyxl"
            ) from exc

        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = options["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Sheet {sheet_name!r} not found. Available: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]

        all_rows = list(worksheet.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError("The worksheet is empty.")

        header = [(c or "").strip() if isinstance(c, str) else c for c in all_rows[0]]
        header = [h for h in header if h is not None]
        missing = [h for h in EXPECTED_HEADERS if h not in header]
        if missing:
            raise CommandError(
                "Workbook headers do not match the expected dataset. "
                f"Missing columns: {missing}"
            )

        index_by_header = {h: i for i, h in enumerate(all_rows[0]) if isinstance(h, str)}
        row_dicts: list[dict] = []
        for raw in all_rows[1:]:
            if not raw or all(cell in (None, "") for cell in raw):
                continue
            row_dicts.append(
                {h: (raw[i] if i < len(raw) else None) for h, i in index_by_header.items()}
            )

        self.stdout.write(
            f"Loaded {len(row_dicts)} data rows from {path.name} (sheet {sheet_name!r})."
        )

        dry_run = options["dry_run"]
        import_kwargs = dict(
            replace_existing=options["replace_existing"],
            include_questionable=options["include_questionable_stats"],
            source_label=options["source_label"],
            default_verification=options["default_verification"],
        )

        if dry_run:
            with transaction.atomic():
                report = import_rows(row_dicts, **import_kwargs)
                transaction.set_rollback(True)
        else:
            with transaction.atomic():
                report = import_rows(row_dicts, **import_kwargs)

        summary = report.as_dict()["summary"]
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Import summary"))
        self.stdout.write(f"  mode:                 {'DRY RUN (rolled back)' if dry_run else 'committed'}")
        self.stdout.write(f"  universities created: {summary['created']}")
        self.stdout.write(f"  universities updated: {summary['updated']}")
        self.stdout.write(f"  rows skipped:         {summary['skipped']}")
        self.stdout.write(f"  warnings:             {summary['warnings']}")
        self.stdout.write(f"  placeholder SAT rows: {summary['placeholder_sat']}")
        self.stdout.write(f"  parsed deadlines:     {summary['parsed_deadlines']}")
        self.stdout.write(f"  parsed essays:        {summary['parsed_essays']}")
        self.stdout.write(f"  source URLs imported: {summary['source_urls']}")
        self.stdout.write(f"  field verifications:  {summary['fields_verified']}")

        report_path = Path(options["report"]) if options["report"] else None
        if report_path is None and not dry_run:
            stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
            report_path = path.parent / f"import_report_{stamp}.json"
        if report_path is not None:
            report_path.write_text(
                json.dumps(report.as_dict(), indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            self.stdout.write(f"  report written to:    {report_path}")

        skipped_rows = [r for r in report.rows if r.status == "skipped"]
        if skipped_rows:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Skipped rows:"))
            for row in skipped_rows:
                self.stdout.write(f"  row {row.row_number}: {row.name} — {'; '.join(row.warnings)}")
