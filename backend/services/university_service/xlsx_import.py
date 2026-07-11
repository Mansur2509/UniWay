"""Rule-based importer for the university XLSX dataset.

The parsing/normalization logic lives here (not in the management command) so it
can be unit-tested directly against in-memory row dicts without touching a file.

Policy (see docs/DECISIONS.md): never invent data. Anything that cannot be
parsed confidently is preserved as raw text and the row is flagged in the import
report. Questionable values (e.g. identical SAT 25/50/75 placeholders) are never
stored as verified statistics.
"""

from __future__ import annotations

import logging
import re
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils.text import slugify

from .currency import normalize_university_costs
from .models import (
    University,
    UniversityDataSource,
    UniversityFieldVerification,
    UniversityProgram,
    UniversityScholarship,
)

logger = logging.getLogger(__name__)

EXPECTED_HEADERS = [
    "Name",
    "Country",
    "City",
    "Official Website",
    "Admissions URL",
    "Majors",
    "Deadlines",
    "SAT 25th",
    "SAT 50th",
    "SAT 75th",
    "IELTS Minimum",
    "IELTS Competitive",
    "Average GPA",
    "Acceptance Rate",
    "Tuition",
    "Scholarships",
    "AP Recommendations by Major",
    "Application Requirements",
    "Essays",
    "Financial Aid",
    "Source URLs",
    "Last Verified Date",
]

DEFAULT_SHEET_NAME = "Database"

VERIFICATION_CHOICES = {"verified", "partial", "estimated"}

# Excel stores dates as a serial day count from 1899-12-30 (the well-known Lotus
# 1900 leap-year quirk offset).
_EXCEL_EPOCH = date(1899, 12, 30)

_URL_RE = re.compile(r"https?://[^\s,;]+", re.IGNORECASE)
_DEADLINE_DATE_RE = re.compile(r"([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})")
_WORD_LIMIT_RE = re.compile(r"([\d,]{2,6})\s*(?:words|characters)", re.IGNORECASE)

_CURRENCY_BY_COUNTRY = {
    "usa": "USD",
    "united states": "USD",
    "uk": "GBP",
    "united kingdom": "GBP",
    "england": "GBP",
    "scotland": "GBP",
    "italy": "EUR",
    "france": "EUR",
    "germany": "EUR",
    "netherlands": "EUR",
    "spain": "EUR",
    "ireland": "EUR",
    "singapore": "SGD",
    "japan": "JPY",
    "china": "CNY",
    "hong kong": "HKD",
    "canada": "CAD",
    "switzerland": "CHF",
    "australia": "AUD",
    "south korea": "KRW",
}

_SCHOLARSHIP_TYPES = [
    ("full ride", "Full-ride scholarship"),
    ("full tuition", "Full-tuition scholarship"),
    ("need-based", "Need-based aid"),
    ("need based", "Need-based aid"),
    ("merit", "Merit scholarship"),
    ("partial", "Partial scholarship"),
    ("international", "International-student aid"),
    ("external", "External scholarship"),
    ("government", "Government/national scholarship"),
    ("national scholarship", "Government/national scholarship"),
]


class UniversityWorkbookError(ValueError):
    """Raised when the XLSX workbook cannot be safely parsed for import."""


def load_xlsx_rows(path: str | Path, *, sheet_name: str = DEFAULT_SHEET_NAME) -> list[dict]:
    """Read the expected university workbook into importer row dictionaries."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise UniversityWorkbookError(
            "openpyxl is required to read university XLSX workbooks."
        ) from exc

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise UniversityWorkbookError(f"Workbook not found: {workbook_path}")

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise UniversityWorkbookError("The uploaded file could not be read as an XLSX workbook.") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise UniversityWorkbookError(
                f"Sheet {sheet_name!r} not found. Available: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]

        all_rows = list(worksheet.iter_rows(values_only=True))
        if not all_rows:
            raise UniversityWorkbookError("The worksheet is empty.")

        header = [(c or "").strip() if isinstance(c, str) else c for c in all_rows[0]]
        header = [h for h in header if h is not None]
        missing = [h for h in EXPECTED_HEADERS if h not in header]
        if missing:
            raise UniversityWorkbookError(
                "Workbook headers do not match the expected dataset. "
                f"Missing columns: {missing}"
            )

        index_by_header = {h: i for i, h in enumerate(all_rows[0]) if isinstance(h, str)}
        row_dicts: list[dict] = []
        for raw in all_rows[1:]:
            if not raw or all(cell in (None, "") for cell in raw):
                continue
            row_dicts.append(
                {h: (raw[i] if i < len(raw) else None) for h, i in index_by_header.items()}
            )
        return row_dicts
    finally:
        workbook.close()


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def strip_parenthetical(name: str) -> str:
    """Drop a trailing "(...)" so "MIT (MIT)" and a curated "MIT" share a slug."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()


def make_slug(name: str) -> str:
    base = strip_parenthetical(name) or name
    return slugify(base)[:260]


def parse_decimal(value) -> Decimal | None:
    s = clean(value).replace(",", "")
    if not s:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def parse_int(value) -> int | None:
    parsed = parse_decimal(value)
    return int(parsed) if parsed is not None else None


def parse_acceptance_rate(value) -> tuple[Decimal | None, str | None]:
    """Return (percent number rounded to 2dp, warning).

    Accepts 0.038 (fraction), "28.0%", "0.28", or 28.0 and always returns a
    percentage number consistent with the existing catalog (e.g. MIT 4.60).
    """
    s = clean(value)
    if not s:
        return None, None
    number = parse_decimal(s)
    if number is None:
        return None, f"unparseable acceptance rate: {s!r}"
    if "%" in s:
        percent = number
    elif number <= 1:
        percent = number * 100
    else:
        percent = number
    percent = percent.quantize(Decimal("0.01"))
    if percent < 0 or percent > 100:
        return None, f"acceptance rate out of range: {s!r}"
    return percent, None


def parse_gpa_average(value) -> tuple[Decimal | None, Decimal | None, str | None]:
    """Return (value, scale, warning) for Average GPA.

    Keeps the legacy unlabeled <=4.50 catalogue convention, but preserves
    explicit source scales such as 88/100 or 4.9/5 so fit comparisons can use
    normalized percentages instead of raw mismatched values.
    """

    s = clean(value)
    if not s:
        return None, None, None
    numbers = re.findall(r"-?\d+(?:\.\d+)?", s.replace(",", ""))
    if not numbers:
        return None, None, f"GPA (as provided): {s}"
    try:
        gpa = Decimal(numbers[0]).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None, None, f"GPA (as provided): {s}"

    scale = None
    if len(numbers) >= 2:
        try:
            candidate_scale = Decimal(numbers[1]).quantize(Decimal("0.01"))
        except InvalidOperation:
            candidate_scale = None
        if candidate_scale in {
            Decimal("4.00"),
            Decimal("5.00"),
            Decimal("10.00"),
            Decimal("20.00"),
            Decimal("45.00"),
            Decimal("100.00"),
        }:
            scale = candidate_scale
        else:
            return None, None, f"unsupported GPA scale: {s!r}"
    elif "%" in s or "percent" in s.lower() or "percentage" in s.lower():
        scale = Decimal("100.00")
    elif gpa <= Decimal("4.50"):
        scale = None
    else:
        return None, None, f"GPA scale not explicit: {s!r}"

    if gpa < 0 or (scale is not None and gpa > scale):
        return None, None, f"GPA out of range: {s!r}"
    return gpa, scale, None


def guess_currency(country: str, raw: str) -> str:
    if "£" in raw:
        return "GBP"
    if "€" in raw:
        return "EUR"
    if "¥" in raw:
        return "JPY"
    lowered = (country or "").lower()
    if "$" in raw and "singapore" in lowered:
        return "SGD"
    if "$" in raw:
        return "USD"
    for key, code in _CURRENCY_BY_COUNTRY.items():
        if key in lowered:
            return code
    return "USD"


def parse_tuition(value, country: str) -> tuple[Decimal | None, str, str, str | None]:
    """Return (amount, currency, raw_text, warning)."""
    s = clean(value)
    if not s:
        return None, "USD", "", None
    currency = guess_currency(country, s)
    number = parse_decimal(s)
    if number is None:
        return None, currency, s, f"unparseable tuition, kept raw: {s!r}"
    return number.quantize(Decimal("0.01")), currency, s, None


def parse_date(value) -> tuple[date | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    s = clean(value)
    if not s:
        return None, None
    if re.fullmatch(r"\d{4,6}(?:\.0)?", s):
        try:
            serial = int(float(s))
        except ValueError:
            serial = None
        if serial is not None and 20000 < serial < 60000:
            return _EXCEL_EPOCH + timedelta(days=serial), None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            continue
    return None, f"unsupported date format: {s!r}"


def detect_sat_placeholder(p25: int | None, p50: int | None, p75: int | None) -> bool:
    """Real SAT 25/50/75 percentiles are never identical; equal values across at
    least two percentiles signal placeholder data (e.g. the 550/550/550 rows)."""
    values = [v for v in (p25, p50, p75) if v is not None]
    return len(values) >= 2 and len(set(values)) == 1


def parse_majors(value) -> list[str]:
    s = clean(value)
    if not s:
        return []
    parts = re.split(r"[;,]", s)
    seen: list[str] = []
    for part in parts:
        name = part.strip()
        if name and name.lower() not in {p.lower() for p in seen}:
            seen.append(name[:240])
    return seen


def parse_sources(value) -> list[str]:
    s = clean(value)
    if not s:
        return []
    urls: list[str] = []
    for match in _URL_RE.findall(s):
        url = match.rstrip(").,;")
        if url not in urls:
            urls.append(url)
    return urls


def parse_deadlines(value) -> list[tuple[str, date | None, str]]:
    s = clean(value)
    if not s:
        return []
    parsed: list[tuple[str, date | None, str]] = []
    for line in re.split(r"[\r\n]+", s):
        line = line.strip()
        if not line:
            continue
        deadline: date | None = None
        match = _DEADLINE_DATE_RE.search(line)
        if match:
            month, day, year = match.group(1), match.group(2), match.group(3)
            for fmt in ("%b %d %Y", "%B %d %Y"):
                try:
                    deadline = datetime.strptime(
                        f"{month[:9]} {day} {year}", fmt
                    ).date()
                    break
                except ValueError:
                    continue
            if deadline is None:
                try:
                    deadline = datetime.strptime(
                        f"{month[:3]} {day} {year}", "%b %d %Y"
                    ).date()
                except ValueError:
                    deadline = None
        label = line.split(":", 1)[0].strip() if ":" in line else line
        parsed.append((label, deadline, line))
    return parsed


def choose_application_deadline(value) -> tuple[date | None, int]:
    parsed = parse_deadlines(value)
    dated = [(label, deadline) for (label, deadline, _raw) in parsed if deadline]
    count = len(dated)
    if not dated:
        return None, count
    for key in ("regular decision", "regular", "application deadline", "ucas", "rd"):
        for label, deadline in dated:
            if key in label.lower():
                return deadline, count
    return max(deadline for _label, deadline in dated), count


def count_word_limits(value) -> int:
    return len(_WORD_LIMIT_RE.findall(clean(value)))


def detect_scholarship_types(value) -> list[str]:
    lowered = clean(value).lower()
    if not lowered:
        return []
    labels: list[str] = []
    for keyword, label in _SCHOLARSHIP_TYPES:
        if keyword in lowered and label not in labels:
            labels.append(label)
    return labels


@dataclass
class RowResult:
    row_number: int
    name: str
    slug: str
    status: str = "skipped"  # created | updated | skipped
    duplicate_matched: bool = False
    parsed_field_count: int = 0
    source_url_count: int = 0
    last_verified_date: str | None = None
    warnings: list[str] = field(default_factory=list)
    questionable_fields: list[str] = field(default_factory=list)


@dataclass
class ImportReport:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    placeholder_sat: int = 0
    parsed_deadlines: int = 0
    parsed_essays: int = 0
    source_urls: int = 0
    fields_verified: int = 0
    rows: list[RowResult] = field(default_factory=list)

    @property
    def warnings_count(self) -> int:
        return sum(len(row.warnings) for row in self.rows)

    def add(self, row: RowResult) -> None:
        self.rows.append(row)
        if row.status == "created":
            self.created += 1
        elif row.status == "updated":
            self.updated += 1
        else:
            self.skipped += 1

    def as_dict(self) -> dict:
        return {
            "summary": {
                "created": self.created,
                "updated": self.updated,
                "skipped": self.skipped,
                "warnings": self.warnings_count,
                "placeholder_sat": self.placeholder_sat,
                "parsed_deadlines": self.parsed_deadlines,
                "parsed_essays": self.parsed_essays,
                "source_urls": self.source_urls,
                "fields_verified": self.fields_verified,
            },
            "rows": [
                {
                    "row_number": row.row_number,
                    "name": row.name,
                    "slug": row.slug,
                    "status": row.status,
                    "duplicate_matched": row.duplicate_matched,
                    "parsed_field_count": row.parsed_field_count,
                    "questionable_fields": row.questionable_fields,
                    "source_url_count": row.source_url_count,
                    "last_verified_date": row.last_verified_date,
                    "warnings": row.warnings,
                }
                for row in self.rows
            ],
        }


def _http(url: str) -> str:
    url = (url or "").strip()
    return url if url.lower().startswith(("http://", "https://")) else ""


@dataclass
class ParsedRow:
    """Pure, DB-free result of parsing one workbook row.

    Both the read-only planner and the writer build from this — no parsing logic
    is duplicated, and parsing never touches the database.
    """

    index: int
    name: str
    slug: str
    valid: bool
    warnings: list[str] = field(default_factory=list)
    questionable_fields: list[str] = field(default_factory=list)

    official: str = ""
    field_values: dict = field(default_factory=dict)
    tuition: Decimal | None = None
    currency: str = "USD"
    scholarship_available: bool | None = None

    sources: list[str] = field(default_factory=list)
    primary_source: str = ""
    verified_date: date | None = None
    last_verified_iso: str | None = None
    majors: list[str] = field(default_factory=list)
    scholarship_labels: list[str] = field(default_factory=list)
    scholarships_text: str = ""

    placeholder: bool = False
    deadline_count: int = 0
    essay_has_limits: bool = False
    verifiable: dict = field(default_factory=dict)
    verification_status: str | None = None
    parsed_field_count: int = 0

    @property
    def source_url_count(self) -> int:
        return len(self.sources)

    @property
    def verification_count(self) -> int:
        if not (self.primary_source and self.verified_date):
            return 0
        return sum(1 for value in self.verifiable.values() if value not in (None, ""))


ImportProgressCallback = Callable[[ImportReport, ParsedRow, int], None]


def parse_row(raw_row: dict, index: int, *, include_questionable: bool, default_verification: str) -> ParsedRow | None:
    """Parse and normalize a single row. Pure: performs no database access.

    Returns ``None`` for an empty (nameless) row, a ``ParsedRow`` with
    ``valid=False`` for a row that cannot be written (no official URL), or a fully
    populated ``ParsedRow`` otherwise.
    """
    name = clean(raw_row.get("Name"))
    if not name:
        return None

    slug = make_slug(name)
    country = clean(raw_row.get("Country"))
    official = _http(raw_row.get("Official Website")) or _http(raw_row.get("Admissions URL"))
    if not official:
        return ParsedRow(
            index=index,
            name=name,
            slug=slug,
            valid=False,
            warnings=["missing official website / admissions URL"],
        )

    warnings: list[str] = []
    questionable_fields: list[str] = []

    sources = parse_sources(raw_row.get("Source URLs"))
    primary_source = _http(sources[0]) if sources else ""
    verified_date, date_warning = parse_date(raw_row.get("Last Verified Date"))
    if date_warning:
        warnings.append(date_warning)
    if not sources:
        warnings.append("missing source URL")

    verification_status = default_verification if (primary_source and verified_date) else None

    acceptance, acceptance_warning = parse_acceptance_rate(raw_row.get("Acceptance Rate"))
    if acceptance_warning:
        warnings.append(acceptance_warning)

    sat25 = parse_int(raw_row.get("SAT 25th"))
    sat50 = parse_int(raw_row.get("SAT 50th"))
    sat75 = parse_int(raw_row.get("SAT 75th"))
    placeholder = detect_sat_placeholder(sat25, sat50, sat75)
    sat_note = ""
    if placeholder:
        warnings.append("possible placeholder SAT values")
        questionable_fields.append("sat")
        sat_note = (
            "Possible placeholder SAT values "
            f"({sat25}/{sat50}/{sat75}); not stored as verified statistics."
        )
        if not include_questionable:
            sat25 = sat50 = sat75 = None

    gpa, gpa_scale, gpa_note = parse_gpa_average(raw_row.get("Average GPA"))
    if gpa_note:
        warnings.append("GPA stored as text note")

    ielts_min = parse_decimal(raw_row.get("IELTS Minimum"))
    ielts_comp = parse_decimal(raw_row.get("IELTS Competitive"))

    tuition, currency, tuition_raw, tuition_warning = parse_tuition(raw_row.get("Tuition"), country)
    if tuition_warning:
        warnings.append(tuition_warning)

    app_deadline, deadline_count = choose_application_deadline(raw_row.get("Deadlines"))
    deadlines_text = clean(raw_row.get("Deadlines"))
    if deadlines_text and deadline_count == 0:
        warnings.append("unparseable deadline (kept raw text)")

    essays_text = clean(raw_row.get("Essays"))
    essay_has_limits = bool(count_word_limits(essays_text))

    majors = parse_majors(raw_row.get("Majors"))
    if not majors:
        warnings.append("empty majors")

    scholarships_text = clean(raw_row.get("Scholarships"))
    financial_aid_notes = clean(raw_row.get("Financial Aid"))
    application_requirements = clean(raw_row.get("Application Requirements"))
    ap_recommendations = clean(raw_row.get("AP Recommendations by Major"))

    data_quality_lines: list[str] = []
    if sat_note:
        data_quality_lines.append(sat_note)
    if gpa_note:
        data_quality_lines.append(gpa_note)
    if tuition is None and tuition_raw:
        data_quality_lines.append(f"Tuition (as provided): {tuition_raw}")
    elif (
        tuition is not None
        and currency == "USD"
        and "$" in tuition_raw
        and any(key in country.lower() for key in _CURRENCY_BY_COUNTRY if _CURRENCY_BY_COUNTRY[key] != "USD")
    ):
        data_quality_lines.append(
            f"Tuition shown as {tuition_raw} in the source (a $ figure on a "
            "non-US institution; treat the currency as the source's USD-equivalent)."
        )
    if not sources:
        data_quality_lines.append(
            "No source URL provided in the dataset; treat all figures as unverified."
        )
    data_quality_notes = "\n".join(data_quality_lines)

    scholarship_available = True if (scholarships_text or financial_aid_notes) else None

    field_values = {
        "country": country,
        "city": clean(raw_row.get("City")),
        "admissions_url": _http(raw_row.get("Admissions URL")),
        "acceptance_rate": acceptance,
        "sat_p25": sat25,
        "sat_average": sat50,
        "sat_p75": sat75,
        "gpa_average": gpa,
        "gpa_average_scale": gpa_scale,
        "ielts_minimum": ielts_min,
        "ielts_competitive": ielts_comp,
        "tuition_amount": tuition,
        "tuition_original_amount": tuition,
        "tuition_original_currency": currency if tuition is not None else "",
        "application_deadline": app_deadline,
        "essay_requirements": essays_text,
        "application_requirements": application_requirements,
        "ap_recommendations": ap_recommendations,
        "deadlines_text": deadlines_text,
        "financial_aid_notes": financial_aid_notes,
        "scholarships_text": scholarships_text,
        "data_quality_notes": data_quality_notes,
    }

    verifiable = {
        "acceptance_rate": acceptance,
        "sat_p25": sat25,
        "sat_average": sat50,
        "sat_p75": sat75,
        "gpa_average": gpa,
        "ielts_minimum": ielts_min,
        "ielts_competitive": ielts_comp,
        "tuition_amount": tuition,
        "application_deadline": app_deadline,
        "essay_requirements": essays_text or None,
        "scholarship_available": scholarship_available,
    }

    parsed_field_count = sum(
        1
        for value in (
            acceptance, sat25, sat50, sat75, gpa, ielts_min, ielts_comp, tuition,
            app_deadline, essays_text or None, scholarships_text or None,
            financial_aid_notes or None, application_requirements or None,
            ap_recommendations or None,
        )
        if value not in (None, "")
    )

    return ParsedRow(
        index=index,
        name=name,
        slug=slug,
        valid=True,
        warnings=warnings,
        questionable_fields=questionable_fields,
        official=official,
        field_values=field_values,
        tuition=tuition,
        currency=currency,
        scholarship_available=scholarship_available,
        sources=sources,
        primary_source=primary_source,
        verified_date=verified_date,
        last_verified_iso=verified_date.isoformat() if verified_date else None,
        majors=majors,
        scholarship_labels=detect_scholarship_types(scholarships_text),
        scholarships_text=scholarships_text,
        placeholder=placeholder,
        deadline_count=deadline_count,
        essay_has_limits=essay_has_limits,
        verifiable=verifiable,
        verification_status=verification_status,
        parsed_field_count=parsed_field_count,
    )


def _accumulate_counts(report: ImportReport, parsed: ParsedRow) -> None:
    report.source_urls += parsed.source_url_count
    report.placeholder_sat += 1 if parsed.placeholder else 0
    report.parsed_deadlines += parsed.deadline_count
    report.parsed_essays += 1 if parsed.essay_has_limits else 0
    report.fields_verified += parsed.verification_count


def _result_from_parsed(parsed: ParsedRow, status: str) -> RowResult:
    return RowResult(
        row_number=parsed.index,
        name=parsed.name,
        slug=parsed.slug,
        status=status,
        duplicate_matched=status == "updated",
        parsed_field_count=parsed.parsed_field_count,
        source_url_count=parsed.source_url_count,
        last_verified_date=parsed.last_verified_iso,
        warnings=list(parsed.warnings),
        questionable_fields=list(parsed.questionable_fields),
    )


def plan_import_rows(
    rows: list[dict],
    *,
    replace_existing: bool = False,  # noqa: ARG001 - accepted for a uniform signature
    include_questionable: bool = False,
    source_label: str = "Universities Data XLSX",  # noqa: ARG001
    default_verification: str = "partial",
) -> ImportReport:
    """TRUE read-only preflight.

    Parses every row and decides created/updated/skipped using a single bulk
    ``slug__in`` SELECT. It never calls ``save()``, ``get_or_create()``,
    ``update_or_create()``, never locks rows, and never relies on rollback — so it
    cannot time out while locking ``university_service_university``.
    """
    if default_verification not in VERIFICATION_CHOICES:
        default_verification = "partial"
    report = ImportReport()

    parsed_rows: list[ParsedRow] = []
    for index, raw_row in enumerate(rows, start=2):  # row 1 is the header
        parsed = parse_row(
            raw_row, index, include_questionable=include_questionable, default_verification=default_verification
        )
        if parsed is not None:
            parsed_rows.append(parsed)

    slugs = [parsed.slug for parsed in parsed_rows if parsed.valid]
    existing_slugs: set[str] = set()
    if slugs:
        # The only database access in dry-run: a single read-only SELECT.
        existing_slugs = set(
            University.objects.filter(slug__in=slugs).values_list("slug", flat=True)
        )

    for parsed in parsed_rows:
        if not parsed.valid:
            report.add(_result_from_parsed(parsed, "skipped"))
            continue
        status = "updated" if parsed.slug in existing_slugs else "created"
        _accumulate_counts(report, parsed)
        report.add(_result_from_parsed(parsed, status))

    return report


# Backwards-compatible alias for the read-only preflight.
dry_run_import_rows = plan_import_rows


def _write_parsed_row(parsed: ParsedRow, *, replace_existing: bool, source_label: str) -> bool:
    """Write one parsed row inside its own short transaction. Returns created?."""
    existing = University.objects.filter(slug=parsed.slug).first()
    created = existing is None
    university = existing or University(
        slug=parsed.slug,
        name=parsed.name,
        official_website=parsed.official,
        is_published=True,
        is_demo=False,
    )

    def assign(attr: str, value) -> None:
        if value in (None, ""):
            return
        if created or replace_existing or getattr(university, attr) in (None, ""):
            setattr(university, attr, value)

    if created:
        university.name = parsed.name
        university.official_website = parsed.official
        university.is_published = True
        university.is_demo = False

    for attr, value in parsed.field_values.items():
        assign(attr, value)

    # Currency has a model default of "USD" (never blank), so tie it to the amount.
    had_tuition = university.tuition_amount is not None
    if parsed.tuition is not None and (created or replace_existing or not had_tuition):
        university.tuition_currency = parsed.currency
        university.tuition_original_amount = parsed.tuition
        university.tuition_original_currency = parsed.currency
        normalize_university_costs(university)

    if parsed.scholarship_available and (
        created or replace_existing or university.scholarship_available is None
    ):
        university.scholarship_available = True

    university.save()

    for major in parsed.majors:
        UniversityProgram.objects.get_or_create(university=university, name=major)

    for url in parsed.sources:
        normalized = _http(url)
        if not normalized:
            continue
        UniversityDataSource.objects.get_or_create(
            university=university,
            source_url=normalized,
            defaults={
                "source_title": f"{source_label}: {parsed.name}",
                "is_official": True,
                "published_at": parsed.verified_date,
            },
        )

    scholarship_url = parsed.primary_source or university.financial_aid_url or parsed.official
    for label in parsed.scholarship_labels:
        UniversityScholarship.objects.get_or_create(
            university=university,
            name=label,
            defaults={"summary": parsed.scholarships_text[:2000], "official_url": scholarship_url},
        )

    if parsed.primary_source and parsed.verified_date:
        for field_name, value in parsed.verifiable.items():
            if value in (None, ""):
                continue
            field_status = parsed.verification_status
            note = ""
            if field_name.startswith("sat") and parsed.placeholder:
                field_status = "estimated"
                note = "Identical SAT percentiles in the source; treat as estimated."
            defaults = {
                "status": field_status,
                "source_url": parsed.primary_source,
                "last_verified_date": parsed.verified_date,
                "note": note,
            }
            if replace_existing:
                UniversityFieldVerification.objects.update_or_create(
                    university=university, field_name=field_name, defaults=defaults
                )
            else:
                UniversityFieldVerification.objects.get_or_create(
                    university=university, field_name=field_name, defaults=defaults
                )

    return created


def _notify_progress(
    callback: ImportProgressCallback | None,
    report: ImportReport,
    parsed: ParsedRow,
) -> None:
    if callback is None:
        return
    callback(report, parsed, len(report.rows))


def execute_import_rows(
    rows: list[dict],
    *,
    replace_existing: bool = False,
    include_questionable: bool = False,
    source_label: str = "Universities Data XLSX",
    default_verification: str = "partial",
    progress_callback: ImportProgressCallback | None = None,
) -> ImportReport:
    """Write rows to the database, one short transaction per university.

    Per-row transactions keep any row lock brief (avoiding the long lock-hold that
    a single giant transaction caused on Supabase) and let a single bad row be
    skipped without aborting the whole import. No rows are ever deleted.
    """
    if default_verification not in VERIFICATION_CHOICES:
        default_verification = "partial"
    report = ImportReport()

    for index, raw_row in enumerate(rows, start=2):
        parsed = parse_row(
            raw_row, index, include_questionable=include_questionable, default_verification=default_verification
        )
        if parsed is None:
            continue
        if not parsed.valid:
            report.add(_result_from_parsed(parsed, "skipped"))
            _notify_progress(progress_callback, report, parsed)
            continue

        try:
            logger.info(
                "University import writing row %s: %s",
                parsed.index,
                parsed.name,
            )
            with transaction.atomic():
                created = _write_parsed_row(parsed, replace_existing=replace_existing, source_label=source_label)
        except Exception as exc:  # noqa: BLE001 - isolate a bad row, never abort the run
            result = _result_from_parsed(parsed, "skipped")
            result.warnings.append(f"row not imported: {exc}")
            report.add(result)
            _notify_progress(progress_callback, report, parsed)
            continue

        _accumulate_counts(report, parsed)
        report.add(_result_from_parsed(parsed, "created" if created else "updated"))
        _notify_progress(progress_callback, report, parsed)

    return report


# Backwards-compatible alias: the historical name now writes via the executor.
import_rows = execute_import_rows
